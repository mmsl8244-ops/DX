import os
import re
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *

#== Local Imports ==
from config_recipe import USER_ID
from utils_recipe import CsvRecipeParser, make_order, parse_order
from ui.widgets import TableViewWithCopyPaste, DraggableTableWidget
from database_service import RecipeService

#font setting
font = QFont()
font.setPixelSize(15)#11pt

class RecipeImportDialog(QDialog):
    """
    다중 CSV 가져오기 설정을 위한 UI를 담당합니다.
    이 클래스는 데이터 처리나 DB 저장 로직을 갖지 않고, 오직 사용자 설정을 수집하여
    상위 위젯(RecipeWindow)에 전달하는 역할만 합니다.
    """

    def __init__(self, parent, chamber_id: str, csv_paths: list[str]):
        super().__init__(parent)
        self._parent_win = parent
        self.chamber_id = chamber_id
        self.csv_paths = csv_paths

        # UI에 표시하기 위해 파싱된 파일의 구조 정보
        self.files_data: list[dict] = []
        # 최종적으로 부모 윈도우에 전달할 설정값
        self._import_configs: list[dict] = []

        self.setWindowTitle(f"Import Recipes for {chamber_id} (Multi-File / Column-Oriented)")
        self.resize(1000, 700)
        self.setFont(font)

        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # 상단 Sync 체크박스
        hl_top = QHBoxLayout()
        label_top = QLabel(f"Latest Recipe: {getattr(self._parent_win, 'latest_recipe_name', '(none)')}", self)

        # [Sync 옵션 설명 수정]
        self.sync_chk = QCheckBox("Sync Settings (Base & Steps follow first column)", self)
        self.sync_chk.setChecked(True)

        hl_top.addWidget(label_top)
        hl_top.addStretch(1)
        hl_top.addWidget(self.sync_chk)
        layout.addLayout(hl_top)

        # CSV 파싱 (UI 구성을 위한 최소한의 정보만)
        parser = CsvRecipeParser()
        for p in self.csv_paths:
            info = parser.parse_file(p)
            if info:
                self.files_data.append(info)

        if not self.files_data:
            QTimer.singleShot(0, lambda: QMessageBox.warning(self, "Error", "No valid CSV files were parsed."))
            QTimer.singleShot(0, self.reject)
            return

        num_files = len(self.files_data)
        max_steps = max(len(inf["steps_info"]) for inf in self.files_data) if self.files_data else 0

        # 테이블 구성
        self.table = QTableWidget(3 + max_steps, 1 + num_files, self)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setEditTriggers(QAbstractItemView.AllEditTriggers)

        # 첫 번째 라벨 컬럼 설정
        self._set_label_cell(0, 0, "File")
        self._set_label_cell(1, 0, "Recipe")
        self._set_label_cell(2, 0, "Base")
        for i in range(max_steps):
            self._set_label_cell(3 + i, 0, f"Step {i + 1}")

        base_candidates = getattr(self._parent_win, "_current_recipe_codes", [])
        latest_base = getattr(self._parent_win, "latest_recipe_name", "")

        # 각 CSV 파일 데이터를 테이블 컬럼으로 채우기
        for c, info in enumerate(self.files_data, start=1):
            # File 행 (수정 불가)
            file_item = QTableWidgetItem(info["recipe_name"])
            file_item.setFlags(file_item.flags() & ~Qt.ItemIsEditable)
            file_item.setData(Qt.UserRole, info["path"])
            self.table.setItem(0, c, file_item)

            # Recipe 행 (수정 가능)
            self.table.setItem(1, c, QTableWidgetItem(info["recipe_name"]))

            # Base 행 (콤보박스)
            combo = QComboBox(self.table)
            combo.addItem("")
            combo.addItems(base_candidates)

            # 기본값 설정
            if latest_base and latest_base in base_candidates:
                combo.setCurrentText(latest_base)

            # [신규] 첫 번째 컬럼(Master)인 경우 시그널 연결
            if c == 1:
                combo.currentTextChanged.connect(self._on_master_base_changed)
            # 나머지 컬럼이고 Sync가 켜져 있으면 비활성화 (초기 상태)
            elif self.sync_chk.isChecked():
                combo.setEnabled(False)

            self.table.setCellWidget(2, c, combo)

            # Step 행들 (체크박스)
            for i in range(max_steps):
                row = 3 + i
                if i < len(info["steps_info"]):
                    step_info = info["steps_info"][i]
                    widget = QWidget(self.table)
                    chk = QCheckBox(step_info["comment"] or "", widget)
                    chk.setChecked(True)
                    chk.setProperty("step_info", step_info)
                    chk.setProperty("file_col", c)
                    chk.setProperty("step_row", row)

                    if c > 1:
                        chk.setEnabled(False)  # Sync 모드가 기본이므로 비활성화
                    else:
                        chk.stateChanged.connect(self._on_master_checkbox_changed)

                    hbox = QHBoxLayout(widget)
                    hbox.setContentsMargins(6, 0, 0, 0)
                    hbox.addWidget(chk)
                    widget.setLayout(hbox)
                    self.table.setCellWidget(row, c, widget)
                else:
                    self.table.setItem(row, c, QTableWidgetItem(""))

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

        # 하단 버튼
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        layout.addWidget(btn_box)

        # 시그널 연결
        btn_box.accepted.connect(self._on_ok)
        btn_box.rejected.connect(self.reject)
        self.sync_chk.stateChanged.connect(self._on_sync_toggle)

    def _set_label_cell(self, row: int, col: int, text: str):
        item = QTableWidgetItem(text)
        item.setFlags(Qt.ItemIsEnabled)
        self.table.setItem(row, col, item)

    def _on_master_base_changed(self, text):
        """[신규] Master Base 콤보박스 변경 시 호출"""
        if not self.sync_chk.isChecked():
            return

        # 나머지 컬럼들의 콤보박스 값을 Master와 동일하게 변경
        for c in range(2, self.table.columnCount()):
            widget = self.table.cellWidget(2, c)
            if isinstance(widget, QComboBox):
                widget.blockSignals(True)  # 불필요한 시그널 방지
                widget.setCurrentText(text)
                widget.blockSignals(False)

    def _on_sync_toggle(self, state):
        is_sync_on = (state == Qt.Checked)
        max_steps = max(len(inf["steps_info"]) for inf in self.files_data) if self.files_data else 0

        # [신규] Base 콤보박스 동기화 처리
        if is_sync_on:
            # 1. Master 값 가져오기
            master_base_widget = self.table.cellWidget(2, 1)
            master_base_val = master_base_widget.currentText() if master_base_widget else ""

            # 2. 나머지 컬럼에 적용 및 비활성화
            self._on_master_base_changed(master_base_val)

        # Base 콤보박스 활성/비활성 상태 변경
        for c in range(2, self.table.columnCount()):
            base_widget = self.table.cellWidget(2, c)
            if base_widget:
                base_widget.setEnabled(not is_sync_on)

        # [기존] Step 체크박스 동기화 처리
        if is_sync_on:
            self._copy_master_states_to_others()

        for r in range(3, 3 + max_steps):
            for c in range(2, self.table.columnCount()):
                widget = self.table.cellWidget(r, c)
                if widget:
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setEnabled(not is_sync_on)

    def _copy_master_states_to_others(self):
        max_steps = max(len(inf["steps_info"]) for inf in self.files_data) if self.files_data else 0
        for r in range(3, 3 + max_steps):
            master_widget = self.table.cellWidget(r, 1)
            is_checked = False
            if master_widget:
                master_chk = master_widget.findChild(QCheckBox)
                if master_chk:
                    is_checked = master_chk.isChecked()

            for c in range(2, self.table.columnCount()):
                widget = self.table.cellWidget(r, c)
                if widget:
                    chk = widget.findChild(QCheckBox)
                    if chk:
                        chk.setChecked(is_checked)

    def _on_master_checkbox_changed(self, state: int):
        if self.sync_chk.isChecked():
            sender_chk = self.sender()
            if not isinstance(sender_chk, QCheckBox): return

            row = sender_chk.property("step_row")
            is_checked = sender_chk.isChecked()

            for c in range(2, self.table.columnCount()):
                widget = self.table.cellWidget(row, c)
                if widget:
                    chk = widget.findChild(QCheckBox)
                    if chk:
                        chk.setChecked(is_checked)

    def _collect_selected_steps_for_column(self, file_col: int) -> list[dict]:
        """특정 파일 컬럼에서 체크된 Step 정보를 목록으로 반환합니다."""
        selected = []
        info = self.files_data[file_col - 1]
        max_steps_for_this_file = len(info["steps_info"])

        for i in range(max_steps_for_this_file):
            row = 3 + i
            widget = self.table.cellWidget(row, file_col)
            if not widget: continue

            chk = widget.findChild(QCheckBox)
            if chk and chk.isChecked():
                selected.append(info["steps_info"][i])
        return selected

    def _on_ok(self):
        """
        사용자 선택을 수집하여 self._import_configs를 구성하고 다이얼로그를 닫습니다.
        DB 로직은 전혀 없습니다.
        """
        configs = []
        for c in range(1, self.table.columnCount()):
            recipe_code_item = self.table.item(1, c)
            recipe_code = recipe_code_item.text().strip() if recipe_code_item else ""

            base_combo = self.table.cellWidget(2, c)
            base_recipe = base_combo.currentText().strip() if base_combo else ""

            selected_steps = self._collect_selected_steps_for_column(c)

            if not recipe_code:
                QMessageBox.warning(self, "Validation Error", f"Recipe code cannot be empty in column {c}.")
                return
            if not selected_steps:
                QMessageBox.warning(self, "Validation Error", f"Please select at least one step in column {c}.")
                return

            config = {
                "path": self.files_data[c - 1]["path"],
                "recipe_code": recipe_code,
                "base_recipe": base_recipe or None,
                "selected_steps": selected_steps
            }
            configs.append(config)

        self._import_configs = configs
        self.accept()

    def get_import_configs(self) -> list[dict]:
        """RecipeWindow가 호출하여 사용자의 최종 설정을 가져갑니다."""
        return self._import_configs

class ColumnEditDialog(QDialog):
    def __init__(self, parent, chamber_id: str, definitions: list[dict], *, auto_shrink: bool = True):
        super().__init__(parent)
        self._parent_win = parent
        self.chamber_id = chamber_id
        self.definitions_data = definitions
        self._auto_shrink = auto_shrink

        # 무한 루프 방지용 플래그
        self._is_batch_updating = False

        self.setWindowTitle(f"Edit Columns for {chamber_id}")
        self.resize(600, 400)
        self.setFont(font)

        # ── 테이블 구성 ─────────────────────────────────────────────
        self.table = DraggableTableWidget(self)
        self.table.setFont(font)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Parameter", "Mapping", "Not Use"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)  # 다중 선택 허용
        self.table.setWordWrap(False)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(False)

        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)

        # ── 버튼 행 ─────────────────────────────────────────────────
        self._btn_layout = QHBoxLayout()
        self.save_btn = QPushButton("Save", self)
        self.exit_btn = QPushButton("Exit", self)
        self._btn_layout.addWidget(self.save_btn)
        self._btn_layout.addWidget(self.exit_btn)

        # ── 레이아웃 조합 ───────────────────────────────────────────
        layout = QVBoxLayout(self)
        layout.addWidget(self.table)
        layout.addLayout(self._btn_layout)

        layout.setSizeConstraint(QLayout.SetMinimumSize)
        self.setSizeGripEnabled(True)

        # 시그널
        self.save_btn.clicked.connect(self.accept)
        self.exit_btn.clicked.connect(self.reject)

        # [신규] 아이템 변경 시그널 (일괄 체크용)
        self.table.itemChanged.connect(self._on_item_changed)

        # 데이터 채우기
        self._populate_table()

        self.table.horizontalHeader().sectionResized.connect(self._recalc_later)
        model = self.table.model()
        if model:
            model.rowsInserted.connect(self._recalc_later)
            model.rowsRemoved.connect(self._recalc_later)
        # itemChanged는 위에서 별도로 연결함

        QTimer.singleShot(0, self._fit_dialog_to_table_width)

    def showEvent(self, e):
        super().showEvent(e)
        self._fit_dialog_to_table_width()

    def _recalc_later(self, *args):
        QTimer.singleShot(0, self._fit_dialog_to_table_width)

    def _populate_table(self):
        self.table.blockSignals(True)  # 초기화 중 시그널 차단
        self.table.setRowCount(len(self.definitions_data))
        for r, item_data in enumerate(self.definitions_data):
            # Parameter
            item0 = QTableWidgetItem(item_data.get("name", ""))
            item0.setFlags(item0.flags() & ~Qt.ItemIsEditable)
            item0.setData(Qt.UserRole, item_data.get("pid"))
            self.table.setItem(r, 0, item0)

            # Mapping
            self.table.setItem(r, 1, QTableWidgetItem(item_data.get("mapping", "")))

            # Hide
            item2 = QTableWidgetItem()
            item2.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
            item2.setCheckState(Qt.Checked if item_data.get("hide") else Qt.Unchecked)
            self.table.setItem(r, 2, item2)
        self.table.blockSignals(False)

    def _on_item_changed(self, item):
        """[신규] 체크박스 변경 시 선택된 다른 행들도 일괄 변경"""
        if self._is_batch_updating:
            return

        # Hide 컬럼(인덱스 2)인지 확인
        if item.column() == 2:
            selected_rows = self.table.selectionModel().selectedRows()
            # 현재 변경된 행이 선택된 행들 중에 포함되어 있는지 확인
            if len(selected_rows) > 1:
                current_row = item.row()
                # 선택된 행 목록에 현재 행이 있는지 확인
                if any(idx.row() == current_row for idx in selected_rows):
                    self._is_batch_updating = True
                    new_state = item.checkState()
                    try:
                        for idx in selected_rows:
                            if idx.row() == current_row: continue

                            target_item = self.table.item(idx.row(), 2)
                            if target_item:
                                target_item.setCheckState(new_state)
                    finally:
                        self._is_batch_updating = False

    def _fit_dialog_to_table_width(self):
        self.table.resizeColumnsToContents()
        h_header = self.table.horizontalHeader()
        v_header = self.table.verticalHeader()

        total_cols_w = sum(h_header.sectionSize(i) for i in range(self.table.columnCount()))
        if v_header.isVisible():
            total_cols_w += v_header.width()
        total_cols_w += self.table.frameWidth() * 2

        vbar = self.table.verticalScrollBar()
        if vbar and vbar.isVisible():
            try:
                total_cols_w += vbar.sizeHint().width()
            except Exception:
                total_cols_w += 16

        lm, tm, rm, bm = self.layout().getContentsMargins()
        needed_w_by_table = total_cols_w + lm + rm
        btn_row_w = self._btn_layout.sizeHint().width() + lm + rm
        needed_width = max(needed_w_by_table, btn_row_w)

        screen = self.windowHandle().screen() if self.windowHandle() else QApplication.primaryScreen()
        avail = screen.availableGeometry() if screen else None
        max_w = int(avail.width() * 0.95) if avail else needed_width

        final_width = min(int(needed_width), max_w)
        if needed_width > max_w:
            self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        else:
            self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        if self._auto_shrink:
            if self.width() != final_width:
                self.setMinimumWidth(final_width)
                self.resize(final_width, self.height())
        else:
            if final_width > self.width():
                self.setMinimumWidth(final_width)
                self.resize(final_width, self.height())

    def get_updated_definitions(self) -> list[dict]:
        updated_defs: list[dict] = []
        for row in range(self.table.rowCount()):
            pid = self.table.item(row, 0).data(Qt.UserRole)
            name = self.table.item(row, 0).text()
            mapping = self.table.item(row, 1).text().strip() if self.table.item(row, 1) else ""
            hide = 1 if (self.table.item(row, 2) and self.table.item(row, 2).checkState() == Qt.Checked) else 0
            updated_defs.append({"pid": pid, "name": name, "mapping": mapping, "hide": hide})
        return updated_defs

class NewRecipeDialog(QDialog):
    # Base 콤보박스 선택이 변경되었음을 알리는 커스텀 시그널
    base_recipe_changed = pyqtSignal(str)

    def __init__(self, parent, base_recipe_list: list[str]):
        super().__init__(parent)
        self.setWindowTitle("New Recipe")
        self.resize(400, 500)
        self.setFont(font)

        # --- UI 구성 ---
        vlay = QVBoxLayout(self)
        form = QFormLayout()

        self.date_edit = QDateEdit(QDate.currentDate(), self)
        self.date_edit.setDisplayFormat("yyMMdd")
        self.date_edit.setCalendarPopup(True)
        form.addRow("Date:", self.date_edit)

        self.comment_edit = QLineEdit(self)
        form.addRow("Comment:", self.comment_edit)

        self.base_combo = QComboBox(self)
        self.base_combo.addItems([""] + base_recipe_list)
        form.addRow("Base:", self.base_combo)

        self.name_edit = QLineEdit(self)
        form.addRow("Recipe Name:", self.name_edit)
        vlay.addLayout(form)

        vlay.addWidget(QLabel("Step:"))

        # [수정 1] TableViewWithCopyPaste 사용 (QTableView 상속)
        self.step_table = TableViewWithCopyPaste(self)

        # [수정 2] 모델(QStandardItemModel) 생성 및 연결
        self.step_model = QStandardItemModel(500, 1)  # 500행 1열
        self.step_model.setHorizontalHeaderLabels(["Step Name"])
        self.step_table.setModel(self.step_model)

        # 헤더 스타일링 (QTableView 방식)
        self.step_table.horizontalHeader().setStretchLastSection(True)

        # 행 높이 설정 (초기값)
        row_h = self.step_table.verticalHeader().defaultSectionSize()
        header_h = self.step_table.horizontalHeader().height()
        # View는 setFixedHeight시 내부 내용물 크기를 자동계산 안하므로 적절히 설정
        # 10개 행 정도 보이는 높이로 설정
        self.step_table.setFixedHeight(row_h * 12)

        vlay.addWidget(self.step_table)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        vlay.addWidget(btn_box)

        # --- 시그널 연결 ---
        self.base_combo.currentTextChanged.connect(self.base_recipe_changed.emit)
        btn_box.accepted.connect(self._on_ok)
        btn_box.rejected.connect(self.reject)

    def _on_ok(self):
        """OK 버튼 클릭 시"""
        if not self.name_edit.text().strip():
            QMessageBox.warning(self, "Validation Error", "Please enter a Recipe Name.")
            return
        self.accept()

    def populate_steps(self, steps: list[str]):
        """
        [수정] QTableView는 clearContents()가 없으므로
        Model을 통해 데이터를 비우고 다시 채웁니다.
        """
        # 기존 데이터 지우기 (빈 문자열로 덮어쓰기)
        for r in range(self.step_model.rowCount()):
            self.step_model.setItem(r, 0, QStandardItem(""))

        # 새 데이터 채우기
        for r, step_name in enumerate(steps):
            if r < self.step_model.rowCount():
                self.step_model.setItem(r, 0, QStandardItem(step_name))

    def get_data(self) -> dict:
        """
        [수정] Model에서 데이터를 읽어옵니다.
        """
        steps = []
        for r in range(self.step_model.rowCount()):
            item = self.step_model.item(r, 0)
            if item and item.text().strip():
                steps.append(item.text().strip())

        return {
            "new_code": self.name_edit.text().strip(),
            "created_at": self.date_edit.date().toString("yyyy-MM-dd") + " " + QTime.currentTime().toString("HH:mm:ss"),
            "comment": self.comment_edit.text().strip(),
            "base_code": self.base_combo.currentText().strip() or None,
            "steps": steps
        }

class CopyRecipeDialog(QDialog):
    """레시피 복사를 위한 소스 선택 UI. DB 접근 로직이 없습니다."""
    # 각 콤보박스 선택이 변경되었음을 알리는 시그널
    process_changed = pyqtSignal(str)
    sheet_changed = pyqtSignal(str)
    chamber_changed = pyqtSignal(str)

    def __init__(self, parent, process_list: list[str], initial_selection: dict):
        super().__init__(parent)
        self.setWindowTitle("Copy Recipe (Select Source)")
        self.setModal(True)
        self.setFont(font)

        # --- UI 위젯 생성 ---
        v_layout = QVBoxLayout(self)
        self.process_combo = self._create_combo_row(v_layout, "Process:")
        self.sheet_combo = self._create_combo_row(v_layout, "Sheet:")
        self.chamber_combo = self._create_combo_row(v_layout, "Chamber:")
        self.base_combo = self._create_combo_row(v_layout, "Base Recipe:")

        spin_layout = QHBoxLayout()
        spin_layout.addWidget(QLabel("How many:", self))
        self.count_spin = QSpinBox(self)
        self.count_spin.setRange(1, 999)
        self.count_spin.setValue(1)
        spin_layout.addWidget(self.count_spin, 1)
        v_layout.addLayout(spin_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        v_layout.addWidget(buttons)

        # --- 초기 데이터 채우기 및 시그널 연결 ---
        self.process_combo.addItems(process_list)
        self._set_initial_selection(initial_selection)

        self.process_combo.currentTextChanged.connect(self.process_changed.emit)
        self.sheet_combo.currentTextChanged.connect(self.sheet_changed.emit)
        self.chamber_combo.currentTextChanged.connect(self.chamber_changed.emit)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def _create_combo_row(self, parent_layout: QVBoxLayout, label: str) -> QComboBox:
        """라벨과 콤보박스로 구성된 한 줄(QHBoxLayout)을 생성하고 부모 레이아웃에 추가합니다."""
        layout = QHBoxLayout()
        layout.addWidget(QLabel(label, self))
        combo = QComboBox(self)
        layout.addWidget(combo, 1)
        parent_layout.addLayout(layout)
        return combo

    def _set_initial_selection(self, selection: dict):
        """메인 윈도우의 현재 선택값을 다이얼로그의 기본값으로 설정합니다."""
        if selection.get("process") and self.process_combo.findText(selection["process"]) >= 0:
            self.process_combo.setCurrentText(selection["process"])
        # Sheet, Chamber, Base는 RecipeWindow의 컨트롤러가 채워줄 것임

    @staticmethod
    def _update_combo(combo: QComboBox, items: list[str], default_item: str = None):
        """콤보박스의 내용을 새로운 아이템 목록으로 안전하게 업데이트합니다."""
        combo.blockSignals(True)
        current = combo.currentText()
        combo.clear()
        combo.addItems(items)

        if default_item and default_item in items:
            combo.setCurrentText(default_item)
        elif current in items:
            combo.setCurrentText(current)
        combo.blockSignals(False)
        # 수동으로 시그널 발생시켜 연쇄 업데이트 유도
        combo.currentTextChanged.emit(combo.currentText())

    def update_sheets(self, sheets: list[str], default: str = None):
        self._update_combo(self.sheet_combo, sheets, default)

    def update_chambers(self, chambers: list[str], default: str = None):
        self._update_combo(self.chamber_combo, chambers, default)

    def update_base_recipes(self, recipes: list[str], default: str = None):
        self._update_combo(self.base_combo, recipes, default)

    def get_source_selection(self) -> dict | None:
        """사용자의 최종 선택을 딕셔너리로 반환합니다."""
        selection = {
            "process": self.process_combo.currentText(),
            "sheet": self.sheet_combo.currentText(),
            "chamber": self.chamber_combo.currentText(),
            "base_code": self.base_combo.currentText(),
            "quantity": self.count_spin.value()
        }
        if not all(selection.values()):
            return None
        return selection

class ProEditDialog(QDialog):
    # UI에서 발생한 이벤트를 알리는 시그널
    process_changed = pyqtSignal(str)
    new_process_requested = pyqtSignal()
    delete_process_requested = pyqtSignal(str)

    def __init__(self, parent, initial_data, recipe_service):
        super().__init__(parent)
        self.recipe_service = recipe_service
        self.setWindowTitle("Classification Metadata")
        self.resize(450, 150)
        self.setFont(font)

        form = QFormLayout(self)

        # 1. Chamber
        self.chamber_combo = QComboBox(self)
        self.chamber_combo.addItems(initial_data.get("chambers", []))
        form.addRow("Chamber:", self.chamber_combo)

        # 2. Process
        proc_container = QWidget()
        proc_layout = QHBoxLayout(proc_container)
        proc_layout.setContentsMargins(0, 0, 0, 0)
        self.process_combo = QComboBox()
        self.process_combo.addItems(initial_data.get("processes", []))
        self.new_proc_btn = QPushButton("New")
        self.del_proc_btn = QPushButton("Delete")
        proc_layout.addWidget(self.process_combo, 1)
        proc_layout.addWidget(self.new_proc_btn)
        proc_layout.addWidget(self.del_proc_btn)
        form.addRow("Process:", proc_container)

        # 3. Sheet (Rename 버튼 추가됨)
        sheet_container = QWidget()
        sheet_layout = QHBoxLayout(sheet_container)
        sheet_layout.setContentsMargins(0, 0, 0, 0)
        self.sheet_combo = QComboBox()
        self.new_sheet_btn = QPushButton("New")
        self.rename_sheet_btn = QPushButton("Rename")  # [신규]
        self.del_sheet_btn = QPushButton("Delete")

        sheet_layout.addWidget(self.sheet_combo, 1)
        sheet_layout.addWidget(self.new_sheet_btn)
        sheet_layout.addWidget(self.rename_sheet_btn)  # [신규] 레이아웃 추가
        sheet_layout.addWidget(self.del_sheet_btn)
        form.addRow("Sheet:", sheet_container)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        form.addRow(buttons)

        # --- 시그널 연결 ---
        self.new_proc_btn.clicked.connect(self.new_process_requested.emit)
        self.del_proc_btn.clicked.connect(lambda: self.delete_process_requested.emit(self.process_combo.currentText()))
        self.process_combo.currentTextChanged.connect(self._on_process_changed)
        self.chamber_combo.currentTextChanged.connect(self._refresh_sheets)

        self.new_sheet_btn.clicked.connect(self._on_new_sheet)
        self.del_sheet_btn.clicked.connect(self._on_delete_sheet)

        # [신규] Rename 버튼 연결
        self.rename_sheet_btn.clicked.connect(self._on_rename_sheet)

        buttons.rejected.connect(self.reject)

        if self.process_combo.count() > 0:
            self._on_process_changed(self.process_combo.currentText())

    def update_processes(self, processes: list[str], select_process: str = None):
        self.process_combo.blockSignals(True)
        self.process_combo.clear()
        self.process_combo.addItems(processes)
        if select_process and select_process in processes:
            self.process_combo.setCurrentText(select_process)
        self.process_combo.blockSignals(False)
        # 목록 갱신 후 현재 선택된 프로세스로 Sheet 갱신 트리거
        self._on_process_changed(self.process_combo.currentText())

    def update_scheme_codes(self, schemes: list[str]):
        # NewSheetDialog에 전달하기 위해 저장만 해둠 (UI에는 표시 안 함)
        self._current_schemes = schemes

    def _on_process_changed(self, process_name):
        # 부모(RecipeWindow)에게 알려 Scheme 목록을 업데이트하도록 함 (내부 self._current_schemes 갱신용)
        self.process_changed.emit(process_name)
        # Sheet 목록 갱신
        self._refresh_sheets()

    def _refresh_sheets(self):
        """현재 선택된 Process/Chamber에 맞는 Sheet 목록 로드"""
        process = self.process_combo.currentText()
        chamber = self.chamber_combo.currentText()

        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()

        if process and chamber:
            # Service를 통해 Sheet 목록 가져오기
            sheets = self.recipe_service.get_sheets_for_chamber(process, chamber)
            self.sheet_combo.addItems(sheets)

        self.sheet_combo.blockSignals(False)

    def _on_new_sheet(self):
        process = self.process_combo.currentText()
        chamber = self.chamber_combo.currentText()

        if not process or not chamber:
            QMessageBox.warning(self, "Warning", "Select Process and Chamber first.")
            return

        # 현재 Process의 Scheme 목록 가져오기
        schemes = self.recipe_service.get_scheme_codes_for_process(process)

        # 새 창 띄우기 (Scheme 목록 전달)
        dlg = NewSheetDialog(self, schemes)
        if dlg.exec_() == QDialog.Accepted:
            data = dlg.get_data()
            # DB 생성 요청
            success, msg = self.recipe_service.create_new_sheet(
                process, chamber, data['sheet'], data['scheme'], data['date']
            )

            if success:
                QMessageBox.information(self, "Success", msg)
                # 목록 갱신 및 새 시트 선택
                self._refresh_sheets()
                idx = self.sheet_combo.findText(data['sheet'])
                if idx >= 0: self.sheet_combo.setCurrentIndex(idx)
            else:
                QMessageBox.warning(self, "Error", msg)

    def _on_delete_sheet(self):
        process = self.process_combo.currentText()
        chamber = self.chamber_combo.currentText()
        sheet = self.sheet_combo.currentText()

        if not sheet: return

        # 1차 삭제 시도
        success, msg, need_confirm = self.recipe_service.delete_sheet(process, chamber, sheet)

        if need_confirm:
            # 레시피가 있어서 경고 필요
            reply = QMessageBox.question(
                self, "Confirm Delete",
                f"Sheet '{sheet}' contains recipes.\n"
                "Are you sure you want to delete this sheet and ALL its recipes?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                # 강제 삭제
                success, msg = self.recipe_service.force_delete_sheet(process, chamber, sheet)
                QMessageBox.information(self, "Result", msg)
                self._refresh_sheets()
        else:
            if success:
                QMessageBox.information(self, "Success", msg)
                self._refresh_sheets()
            else:
                QMessageBox.warning(self, "Error", msg)

    # [신규] Rename 핸들러 구현
    def _on_rename_sheet(self):
        process = self.process_combo.currentText()
        chamber = self.chamber_combo.currentText()
        old_sheet = self.sheet_combo.currentText()

        if not process or not chamber or not old_sheet:
            QMessageBox.warning(self, "Warning", "Select a sheet to rename.")
            return

        # 입력창 띄우기
        new_sheet, ok = QInputDialog.getText(self, "Rename Sheet",
                                             f"Enter new name for '{old_sheet}':",
                                             text=old_sheet)

        if ok and new_sheet:
            new_sheet = new_sheet.strip()
            # 서비스 호출
            success, msg = self.recipe_service.rename_sheet(process, chamber, old_sheet, new_sheet)

            if success:
                QMessageBox.information(self, "Success", msg)
                self._refresh_sheets()  # 목록 갱신
                # 변경된 이름 선택해주기
                idx = self.sheet_combo.findText(new_sheet)
                if idx >= 0:
                    self.sheet_combo.setCurrentIndex(idx)
            else:
                QMessageBox.warning(self, "Error", msg)

class NewSheetDialog(QDialog):
    """새로운 Sheet를 생성하기 위한 다이얼로그"""

    def __init__(self, parent, scheme_list):
        super().__init__(parent)

        self.setWindowTitle("Create New Sheet")
        self.resize(350, 200)
        self.setFont(font)

        layout = QFormLayout(self)

        self.edit_sheet = QLineEdit()
        self.combo_scheme = QComboBox()
        self.combo_scheme.addItems(scheme_list)

        self.edit_date = QDateEdit(QDate.currentDate())
        self.edit_date.setDisplayFormat("yyMMdd")
        self.edit_date.setCalendarPopup(True)

        layout.addRow("Sheet Name:", self.edit_sheet)
        layout.addRow("Scheme Code:", self.combo_scheme)
        layout.addRow("Date:", self.edit_date)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self._validate)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _validate(self):
        if not self.edit_sheet.text().strip():
            QMessageBox.warning(self, "Warning", "Sheet name cannot be empty.")
            return
        self.accept()

    def get_data(self):
        return {
            "sheet": self.edit_sheet.text().strip(),
            "scheme": self.combo_scheme.currentText(),
            "date": int(self.edit_date.date().toString("yyMMdd"))
        }

class CommentEditDialog(QDialog):
    """
    긴 Comment 입력을 위한 전용 팝업 다이얼로그
    - Enter: 저장 및 닫기
    - Shift + Enter: 줄바꿈
    """

    def __init__(self, parent, initial_text):
        super().__init__(parent)

        self.setWindowTitle("Edit Comment")
        self.resize(400, 300)  # 넉넉한 크기

        layout = QVBoxLayout(self)

        # 라벨
        layout.addWidget(QLabel("Edit Comment (Shift+Enter for new line, Enter to Save):"))

        # 텍스트 에디터
        self.text_edit = QPlainTextEdit(self)
        self.text_edit.setPlainText(str(initial_text))

        # 커서를 문서 끝으로 이동
        cursor = self.text_edit.textCursor()
        cursor.movePosition(QTextCursor.End)
        self.text_edit.setTextCursor(cursor)

        # 이벤트 필터 설치 (키 입력 처리를 위해)
        self.text_edit.installEventFilter(self)

        layout.addWidget(self.text_edit)

        # 버튼 박스
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        # 다이얼로그 실행 시 텍스트 에디터에 포커스
        self.text_edit.setFocus()

    def eventFilter(self, obj, event):
        if obj is self.text_edit and event.type() == QEvent.KeyPress:
            if event.key() == Qt.Key_Return:
                if event.modifiers() & Qt.ShiftModifier:
                    # Shift + Enter: 줄바꿈 허용 (기본 동작)
                    return False
                else:
                    # Enter 단독: 저장(OK) 처리
                    self.accept()
                    return True
        return super().eventFilter(obj, event)

    def get_text(self):
        return self.text_edit.toPlainText()

class ImportDBDialog(QDialog):
    def __init__(self, parent, chamber_list: list[str], existing_chambers: set[str]):
        super().__init__(parent)
        self.setWindowTitle("Import Parameter Definitions")
        self.resize(480, 220)
        self.setFont(font)

        self._result = {
            "chamber": "",
            "mode": "",
            "path": "",
            "manual_params": [],
        }

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Select Chamber ID:"))

        self.combo = QComboBox(self)

        # [수정] 챔버 리스트를 추가하면서, 이미 등록된 챔버는 파란색으로 표시
        for ch in chamber_list:
            self.combo.addItem(ch)
            if ch in existing_chambers:
                # 해당 아이템의 인덱스
                idx = self.combo.count() - 1
                # 텍스트 색상을 파란색으로 설정
                self.combo.setItemData(idx, QColor("blue"), Qt.ForegroundRole)
                # (선택 사항) 툴팁 추가
                self.combo.setItemData(idx, "Parameters already defined", Qt.ToolTipRole)

        layout.addWidget(self.combo)

        btns = QHBoxLayout()
        self.csv_btn = QPushButton("CSV Load...", self)
        self.manual_btn = QPushButton("Manual", self)
        self.cancel_btn = QPushButton("Cancel", self)
        btns.addWidget(self.csv_btn)
        btns.addWidget(self.manual_btn)
        btns.addStretch()
        btns.addWidget(self.cancel_btn)
        layout.addLayout(btns)

        self.csv_btn.clicked.connect(self._on_csv)
        self.manual_btn.clicked.connect(self._on_manual)
        self.cancel_btn.clicked.connect(self.reject)

    def _require_chamber(self) -> str:
        c = (self.combo.currentText() or "").strip()
        if not c:
            QMessageBox.warning(self, "Warning", "Please select a Chamber ID.")
            return ""
        return c

    def _on_csv(self):
        chamber_id = self._require_chamber()
        if not chamber_id:
            return
        path, _ = QFileDialog.getOpenFileName(self, "Select a CSV file", "", "CSV Files (*.csv)")
        if not path:
            return
        self._result.update({"chamber": chamber_id, "mode": "csv", "path": path, "manual_params": []})
        self.accept()

    def _on_manual(self):
        chamber_id = self._require_chamber()
        if not chamber_id:
            return

        reply = QMessageBox.question(
            self,
            "Manual Import",
            "Manual import prevents file-based recipe import (CSV/XML).\n"
            "Do you still want to proceed with Manual mode?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        editor = ManualImportEditorDialog(self, chamber_id)
        if editor.exec_() != QDialog.Accepted:
            return

        defs = editor.get_params()  # list[{"name","unit"}]
        if not defs:
            QMessageBox.warning(self, "Warning", "No manual parameters were entered.")
            return

        self._result.update({"chamber": chamber_id, "mode": "manual", "path": "", "manual_params": defs})
        self.accept()

    def get_result(self) -> dict:
        return dict(self._result)

class ManualImportEditorDialog(QDialog):
    """
    수동 파라미터 입력 다이얼로그 (단일 컬럼 "Para").
    - 사용자가 'Para' 목록을 직접 입력/붙여넣기
    - 붙여넣기(Ctrl+V) 시 필요한 만큼 행 자동 추가
    - OK 시 list[str] 반환 (CSV의 Para 컬럼과 동일한 의미)
    """

    def __init__(self, parent=None, chamber_id: str = ""):
        super().__init__(parent)
        self.setWindowTitle(f"Manual Import - {chamber_id or 'Chamber'}")
        self.resize(680, 460)
        self.setFont(font)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Enter parameters (one per row). Use Ctrl+V to paste. (200 rows)"))

        # View/Model
        self.table = TableViewWithCopyPaste(self)
        self.model = QStandardItemModel(200, 2, self)
        self.model.setHorizontalHeaderLabels(["Parameter", "Unit"])
        self.table.setModel(self.model)
        self.table.setEditTriggers(QTableView.AllEditTriggers)
        self.table.setSelectionBehavior(QTableView.SelectItems)
        self.table.setSelectionMode(QTableView.ExtendedSelection)

        header = self.table.horizontalHeader()
        # Stretch 대신 Fixed로 두고, 우리가 동일 폭을 직접 강제합니다.
        header.setSectionResizeMode(QHeaderView.Fixed)
        header.setMinimumSectionSize(120)
        header.setStretchLastSection(False)  # 두 컬럼 동일 폭을 유지하기 위해 비활성화

        # 초기 200행 빈 셀 생성
        for r in range(200):
            self.model.setItem(r, 0, QStandardItem(""))
            self.model.setItem(r, 1, QStandardItem(""))

        layout.addWidget(self.table)

        # OK / Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        layout.addWidget(buttons)
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)

        # 최초 표시 전에 한 번 강제
        QTimer.singleShot(0, self._apply_equal_column_widths)

    # ── 컬럼 동일 폭 강제 로직 ─────────────────────────────────────────
    def _apply_equal_column_widths(self):
        """두 컬럼(0,1)을 항상 동일 폭으로 강제한다."""
        hdr = self.table.horizontalHeader()

        # 테이블 뷰포트 실제 가용 폭
        viewport_w = self.table.viewport().width()

        # 세로 스크롤바가 보이면 그만큼 차감
        vbar = self.table.verticalScrollBar()
        vbar_w = vbar.sizeHint().width() if (vbar and vbar.isVisible()) else 0

        # 프레임 여백 보정
        frame = self.table.frameWidth() * 2

        usable = max(0, viewport_w - vbar_w - frame)
        col_w = max(hdr.minimumSectionSize(), usable // 2)

        hdr.resizeSection(0, col_w)
        hdr.resizeSection(1, col_w)

    def resizeEvent(self, e):
        super().resizeEvent(e)
        # 다이얼로그/테이블 리사이즈 때마다 동일 폭 유지
        self._apply_equal_column_widths()

    def showEvent(self, e):
        super().showEvent(e)
        # 표시 직후 한 번 더 보정(스크롤바 가시성 확정 후)
        QTimer.singleShot(0, self._apply_equal_column_widths)

    # ── 데이터 수집/검증 ─────────────────────────────────────────────
    def get_params(self) -> list:
        """
        각 행의 (Parameter, Unit)을 읽어서
        - Parameter 공백/빈칸은 스킵
        - Parameter 중복은 첫 번째만 사용
        반환: [{"name": <Parameter>, "unit": <Unit>}]
        """
        out = []
        seen = set()
        rc = self.model.rowCount()
        for r in range(rc):
            name = (self.model.index(r, 0).data(Qt.DisplayRole) or "").strip()
            unit = (self.model.index(r, 1).data(Qt.DisplayRole) or "").strip()
            if not name:
                continue
            if name in seen:
                continue
            seen.add(name)
            out.append({"name": name, "unit": unit})
        return out

    def _on_accept(self):
        params = self.get_params()
        if not params:
            QMessageBox.warning(self, "Warning", "Please enter at least one parameter.")
            return
        self.accept()

class RampEditDialog(QDialog):
    """Ramp 파라미터의 시작/끝 값을 편집하기 위한 다이얼로그."""

    def __init__(self, parent, ramp_param_name: str, ramp_times: float, target_params: list[dict]):
        super().__init__(parent)
        self.setWindowTitle(f"Edit Ramp: {ramp_param_name}")
        self.setMinimumWidth(400)
        self.setFont(font)

        self.target_params = target_params
        self.widgets = {}

        layout = QVBoxLayout(self)
        form_layout = QFormLayout()

        # Ramp Times 입력 필드
        self.ramp_times_spin = QDoubleSpinBox(self)
        self.ramp_times_spin.setRange(0, 9999)
        self.ramp_times_spin.setValue(ramp_times or 0)
        form_layout.addRow(f"{ramp_param_name}:", self.ramp_times_spin)

        # 각 대상 파라미터에 대한 시작/끝 값 입력 필드
        for param in target_params:
            container = QWidget()
            hbox = QHBoxLayout(container)
            hbox.setContentsMargins(0, 0, 0, 0)
            start_edit = QLineEdit(self)
            start_edit.setPlaceholderText("Start")
            end_edit = QLineEdit(self)
            end_edit.setPlaceholderText("End")

            # 기존 Ramp 값이 있으면 채워넣기
            if isinstance(param['value'], str) and '>' in param['value']:
                parts = param['value'].split('>', 1)
                start_edit.setText(parts[0].strip())
                end_edit.setText(parts[1].strip())
            else:  # Ramp가 아니면 끝 값만 표시
                end_edit.setText(str(param.get('value', '')))

            hbox.addWidget(start_edit)
            hbox.addWidget(QLabel(">"))
            hbox.addWidget(end_edit)

            form_layout.addRow(f"{param['name']}:", container)
            self.widgets[param['pid']] = (start_edit, end_edit)

        layout.addLayout(form_layout)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        layout.addWidget(buttons)

        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

    def get_ramp_data(self) -> dict:
        """사용자 입력을 딕셔너리로 반환합니다."""
        data = {"ramp_times": self.ramp_times_spin.value(), "params": {}}
        for pid, (start_edit, end_edit) in self.widgets.items():
            start_val = start_edit.text().strip()
            end_val = end_edit.text().strip()
            try:
                if start_val and end_val:
                    data["params"][pid] = {"start": float(start_val), "end": float(end_val)}
                elif end_val:  # 끝 값만 있는 경우
                    data["params"][pid] = {"start": None, "end": float(end_val)}
                else:
                    # 둘 다 비어있으면 Ramp 해제 (aux_value=None, value 유지 안 함)
                    data["params"][pid] = {"start": None, "end": None}
            except ValueError:
                continue
        return data

class DynamicStepEditDialog(QDialog):
    """Dynamic Process의 시작 스텝을 선택하기 위한 다이얼로그."""

    def __init__(self, parent, all_steps: list[dict], entry_step_no: int,
                 initial_repeat: int | None = None, initial_start: int | None = None):
        super().__init__(parent)
        self.setWindowTitle("Select Dynamic Process Start Step")
        self.setMinimumWidth(300)
        self.setFont(font)

        # 1) 진입 스텝 이전만 노출
        self._filtered_steps = [s for s in all_steps if s["step_no"] < int(entry_step_no)]

        main = QVBoxLayout(self)

        # 안내
        main.addWidget(QLabel("Select a dynamic start step (only before the entry step):", self))

        # 스텝 리스트
        self.step_list = QListWidget(self)
        for s in self._filtered_steps:
            self.step_list.addItem(f"Step {s['step_no']}: {s['step_name']}")
        main.addWidget(self.step_list)

        # 2) Dynamic Process 정수 입력
        row = QHBoxLayout()
        row.addWidget(QLabel("Dynamic Process :", self))
        self.repeat_edit = QLineEdit(self)
        self.repeat_edit.setPlaceholderText("Integer")
        self.repeat_edit.setValidator(QIntValidator(0, 999999, self))
        if isinstance(initial_repeat, int):
            self.repeat_edit.setText(str(initial_repeat))
        row.addWidget(self.repeat_edit, 1)
        main.addLayout(row)

        # 초기 선택(시작 스텝)
        if isinstance(initial_start, int):
            for i, s in enumerate(self._filtered_steps):
                if s["step_no"] == initial_start:
                    self.step_list.setCurrentRow(i)
                    break

        # 버튼
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        main.addWidget(btns)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)

    def get_results(self) -> tuple[int | None, int | None]:
        """(선택된 시작 스텝 번호 or None, 입력한 반복 횟수 or None)"""
        start_no = None
        item = self.step_list.currentItem()
        if item:
            m = re.match(r"Step (\d+):", item.text())
            if m:
                start_no = int(m.group(1))
        rep = self.repeat_edit.text().strip()
        repeat = int(rep) if rep.isdigit() else None
        return start_no, repeat

class RowBatchUpdateDialog(QDialog):
    """
    행 단위 일괄 변경 다이얼로그
    - 현재 화면에 보이는 파라미터(컬럼) 목록을 체크박스로 표시
    - 배율(Multiplier) 및 소수점(Precision) 설정
    """

    def __init__(self, parent, columns):
        super().__init__(parent)

        self.setWindowTitle("Batch Change (Selected Rows)")
        self.resize(400, 600)
        self.setFont(font)

        self.selected_indices = []  # 선택된 컬럼의 인덱스 리스트
        self.multiplier = 1.0
        self.precision = 2

        layout = QVBoxLayout(self)

        # 1. 파라미터 선택 영역
        grp_cols = QGroupBox("Select Parameters to Update")
        vbox_cols = QVBoxLayout(grp_cols)

        self.list_widget = QListWidget()
        for col_name in columns:
            # 줄바꿈 문자(\n)가 있으면 보기 좋게 한 줄로 변경
            display_text = col_name.replace('\n', ' ')
            item = QListWidgetItem(display_text)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)

        # "Select All" 버튼 추가
        btn_select_all = QPushButton("Select All")
        btn_select_all.clicked.connect(self._select_all)
        vbox_cols.addWidget(btn_select_all)

        vbox_cols.addWidget(self.list_widget)
        layout.addWidget(grp_cols)

        # 2. 설정 영역
        grp_settings = QGroupBox("Settings")
        form_layout = QFormLayout(grp_settings)

        self.spin_mult = QDoubleSpinBox()
        self.spin_mult.setRange(0.0, 100.0)
        self.spin_mult.setDecimals(2)
        self.spin_mult.setSingleStep(0.1)
        self.spin_mult.setValue(1.0)
        form_layout.addRow("Multiplier (x):", self.spin_mult)

        self.spin_prec = QSpinBox()
        self.spin_prec.setRange(0, 10)
        self.spin_prec.setValue(2)
        form_layout.addRow("Decimal Places:", self.spin_prec)

        layout.addWidget(grp_settings)

        # 3. 버튼
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self._on_accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _select_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Checked)

    def _on_accept(self):
        self.selected_indices = []
        for i in range(self.list_widget.count()):
            if self.list_widget.item(i).checkState() == Qt.Checked:
                self.selected_indices.append(i)

        if not self.selected_indices:
            QMessageBox.warning(self, "Warning", "Please select at least one parameter.")
            return

        self.multiplier = self.spin_mult.value()
        self.precision = self.spin_prec.value()
        self.accept()

    def get_data(self):
        return self.selected_indices, self.multiplier, self.precision

class ParamMappingDialog(QDialog):
    def __init__(self, parent, db_only: list[dict], import_only: list[dict]):
        super().__init__(parent)
        self.setWindowTitle("Parameter Mapping (Update Names)")
        self.resize(900, 600)

        # 데이터 저장소
        self.db_only = db_only  # [{'name', 'unit', 'pid'}, ...]
        self.import_only = import_only  # [{'name', 'unit', 'order', 'mapping'}, ...]
        self.mapping_result = []  # [ (db_item, import_item), ... ]

        layout = QVBoxLayout(self)

        # 안내 문구
        info = QLabel("<b>Link renamed parameters manually to preserve recipe data.</b><br>"
                      "Left: Old params (in DB only) | Right: New params (in File only)<br>"
                      " - Linked: Old parameter name will be updated to new name (ID kept).<br>"
                      " - Unlinked Left: Will be deactivated (Hidden).<br>"
                      " - Unlinked Right: Will be created as new ID.")
        layout.addWidget(info)

        # 메인 영역 (3단 구성)
        h_layout = QHBoxLayout()

        # 1. Left (DB Only)
        grp_left = QGroupBox("Old Parameters (In DB)")
        v_left = QVBoxLayout(grp_left)
        self.list_db = QListWidget()
        self.list_db.setSelectionMode(QAbstractItemView.SingleSelection)
        for item in self.db_only:
            self.list_db.addItem(f"{item['name']} [{item['unit']}]")
        v_left.addWidget(self.list_db)
        h_layout.addWidget(grp_left, 1)

        # 2. Center (Actions & Result)
        grp_center = QGroupBox("Mapped (To be Renamed)")
        v_center = QVBoxLayout(grp_center)

        btn_link = QPushButton("Link (Old -> New) >>")
        btn_unlink = QPushButton("<< Unlink")

        self.table_map = QTableWidget(0, 2)
        self.table_map.setHorizontalHeaderLabels(["Old Name", "New Name"])
        self.table_map.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_map.setSelectionBehavior(QAbstractItemView.SelectRows)

        v_center.addWidget(btn_link)
        v_center.addWidget(self.table_map)
        v_center.addWidget(btn_unlink)
        h_layout.addWidget(grp_center, 2)  # 가운데를 좀 더 넓게

        # 3. Right (Import Only)
        grp_right = QGroupBox("New Parameters (In File)")
        v_right = QVBoxLayout(grp_right)
        self.list_new = QListWidget()
        self.list_new.setSelectionMode(QAbstractItemView.SingleSelection)
        for item in self.import_only:
            self.list_new.addItem(f"{item['name']} [{item['unit']}]")
        v_right.addWidget(self.list_new)
        h_layout.addWidget(grp_right, 1)

        layout.addLayout(h_layout)

        # 하단 버튼
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        layout.addWidget(btn_box)

        # 시그널
        btn_link.clicked.connect(self._on_link)
        btn_unlink.clicked.connect(self._on_unlink)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)

    def _on_link(self):
        row_db = self.list_db.currentRow()
        row_new = self.list_new.currentRow()

        if row_db < 0 or row_new < 0:
            return  # 선택 안됨

        # 데이터 이동
        db_item = self.db_only.pop(row_db)
        new_item = self.import_only.pop(row_new)

        # 리스트 위젯 갱신
        self.list_db.takeItem(row_db)
        self.list_new.takeItem(row_new)

        # 매핑 결과 추가
        self.mapping_result.append((db_item, new_item))

        # 테이블 표시
        r = self.table_map.rowCount()
        self.table_map.insertRow(r)
        self.table_map.setItem(r, 0, QTableWidgetItem(db_item['name']))
        self.table_map.setItem(r, 1, QTableWidgetItem(new_item['name']))

    def _on_unlink(self):
        row = self.table_map.currentRow()
        if row < 0: return

        # 데이터 복구
        db_item, new_item = self.mapping_result.pop(row)

        # 리스트로 복귀
        self.db_only.append(db_item)
        self.import_only.append(new_item)

        # 정렬 (선택사항)
        self.db_only.sort(key=lambda x: x['name'])
        self.import_only.sort(key=lambda x: x['name'])

        # UI 리프레시 (간단히 다시 그림)
        self.list_db.clear()
        for item in self.db_only:
            self.list_db.addItem(f"{item['name']} [{item['unit']}]")

        self.list_new.clear()
        for item in self.import_only:
            self.list_new.addItem(f"{item['name']} [{item['unit']}]")

        self.table_map.removeRow(row)

    def get_results(self):
        """
        반환값:
        1. mapped_pairs: [(db_item, new_item), ...] -> Update Name
        2. final_new_items: [new_item, ...] -> Insert New
        3. final_db_remain: [db_item, ...] -> Soft Delete
        """
        return self.mapping_result, self.import_only, self.db_only

class ExcelFilterDialog(QDialog):
    def __init__(self, parent, col_name, unique_values, current_filter):
        super().__init__(parent)
        self.setWindowTitle(f"Filter: {col_name}")
        self.resize(300, 400)
        main_layout = QVBoxLayout(self)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Search...")
        self.search_edit.textChanged.connect(self._filter_list)
        main_layout.addWidget(self.search_edit)

        self.list_widget = QListWidget()
        main_layout.addWidget(self.list_widget)

        self.all_item = QListWidgetItem("(Select All)")
        self.all_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        # 전체 선택 여부 초기화
        is_all_checked = (current_filter is None)
        self.all_item.setCheckState(Qt.Checked if is_all_checked else Qt.Unchecked)
        self.list_widget.addItem(self.all_item)

        self.items = []
        # 정렬하여 표시
        sorted_values = sorted(list(unique_values), key=lambda x: str(x))
        for val in sorted_values:
            item = QListWidgetItem(str(val))
            item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            if is_all_checked or val in current_filter:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)
            self.items.append(item)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        main_layout.addWidget(btn_box)

        self.list_widget.itemChanged.connect(self._on_item_changed)

    def _on_item_changed(self, item):
        if item == self.all_item:
            state = item.checkState()
            self.list_widget.blockSignals(True)
            for i in self.items:
                i.setCheckState(state)
            self.list_widget.blockSignals(False)

    def _filter_list(self, text):
        for i in range(1, self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setHidden(text.lower() not in item.text().lower())

    def get_selected_values(self):
        # 모두 체크되어 있으면 None 반환 (필터 없음)
        all_checked = True
        for item in self.items:
            if item.checkState() != Qt.Checked:
                all_checked = False
                break

        if all_checked: return None

        selected = set()
        for item in self.items:
            if item.checkState() == Qt.Checked:
                selected.add(item.text())
        return selected


class ExportExcelDialog(QDialog):
    """
    Excel Export 다이얼로그.
    Process → Sheet → Chamber 선택 후 전체 Sheet 또는 특정 Recipe를 Excel로 Export.
    """

    def __init__(self, parent, recipe_service: RecipeService, current_context=None):
        super().__init__(parent)
        self.recipe_service = recipe_service
        self._initial_context = current_context or {}
        self.setWindowTitle("Export to Excel")
        self.resize(550, 500)
        self.setFont(font)
        self._build_ui()
        self._connect_signals()
        self._load_processes()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # ── 1. Process / Sheet / Chamber 선택 ──
        form = QFormLayout()
        self.process_combo = QComboBox(self)
        self.sheet_combo = QComboBox(self)
        self.chamber_combo = QComboBox(self)

        form.addRow("Process:", self.process_combo)
        form.addRow("Sheet:", self.sheet_combo)
        form.addRow("Chamber:", self.chamber_combo)
        layout.addLayout(form)

        # ── 2. Export 범위 선택 ──
        grp = QGroupBox("Export Range", self)
        grp_layout = QVBoxLayout(grp)

        self.radio_all = QRadioButton("All Recipes in Sheet", self)
        self.radio_select = QRadioButton("Selected Recipes Only", self)
        self.radio_all.setChecked(True)

        grp_layout.addWidget(self.radio_all)
        grp_layout.addWidget(self.radio_select)

        # Recipe 리스트 (체크박스)
        self.recipe_list = QListWidget(self)
        self.recipe_list.setEnabled(False)
        grp_layout.addWidget(self.recipe_list)
        layout.addWidget(grp)

        # ── 3. 저장 경로 ──
        path_layout = QHBoxLayout()
        self.path_edit = QLineEdit(self)
        self.path_edit.setPlaceholderText("Select save path...")
        self.path_edit.setReadOnly(True)
        self.browse_btn = QPushButton("Browse...", self)
        path_layout.addWidget(self.path_edit)
        path_layout.addWidget(self.browse_btn)
        layout.addLayout(path_layout)

        # 파일이름 미리보기
        self.filename_label = QLabel("", self)
        self.filename_label.setStyleSheet("color: gray;")
        layout.addWidget(self.filename_label)

        # ── 4. 버튼 ──
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        self.export_btn = QPushButton("Export", self)
        self.export_btn.setEnabled(False)
        self.cancel_btn = QPushButton("Cancel", self)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addWidget(self.cancel_btn)
        layout.addLayout(btn_layout)

    def _connect_signals(self):
        self.process_combo.currentIndexChanged.connect(self._on_process_changed)
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)
        self.chamber_combo.currentIndexChanged.connect(self._on_chamber_changed)
        self.radio_select.toggled.connect(self._on_range_toggled)
        self.browse_btn.clicked.connect(self._browse_path)
        self.export_btn.clicked.connect(self._do_export)
        self.cancel_btn.clicked.connect(self.reject)

    def _load_processes(self):
        processes = self.recipe_service.get_available_processes(only_with_recipes=True)
        self.process_combo.clear()
        self.process_combo.addItems(processes)
        # 현재 컨텍스트가 있으면 자동 선택
        ctx = self._initial_context
        if ctx.get("process"):
            idx = self.process_combo.findText(ctx["process"])
            if idx >= 0:
                self.process_combo.setCurrentIndex(idx)
                # process 변경 시그널로 sheet가 로드된 후 sheet 선택
                if ctx.get("sheet"):
                    idx_s = self.sheet_combo.findText(ctx["sheet"])
                    if idx_s >= 0:
                        self.sheet_combo.setCurrentIndex(idx_s)
                        # sheet 변경 시그널로 chamber가 로드된 후 chamber 선택
                        if ctx.get("chamber"):
                            idx_c = self.chamber_combo.findText(ctx["chamber"])
                            if idx_c >= 0:
                                self.chamber_combo.setCurrentIndex(idx_c)

    def _on_process_changed(self):
        self.sheet_combo.clear()
        self.chamber_combo.clear()
        self.recipe_list.clear()
        process = self.process_combo.currentText().strip()
        if not process:
            return
        sheets = self.recipe_service.get_sheets_for_process(process)
        self.sheet_combo.addItems(sheets)
        self._update_filename()

    def _on_sheet_changed(self):
        self.chamber_combo.clear()
        process = self.process_combo.currentText().strip()
        sheet = self.sheet_combo.currentText().strip()
        if not process or not sheet:
            return
        chambers = self.recipe_service.get_chambers_for_sheet(process, sheet)
        self.chamber_combo.addItems(chambers)
        self._update_filename()

    def _on_chamber_changed(self):
        self._load_recipe_list()
        self._update_filename()
        self._check_export_ready()

    def _load_recipe_list(self):
        self.recipe_list.clear()
        process = self.process_combo.currentText().strip()
        sheet = self.sheet_combo.currentText().strip()
        chamber = self.chamber_combo.currentText().strip()
        if not all([process, sheet, chamber]):
            return
        codes = self.recipe_service.get_recipes_for_chamber(process, sheet, chamber)
        for code in codes:
            item = QListWidgetItem(code)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked)
            self.recipe_list.addItem(item)

    def _on_range_toggled(self, checked):
        self.recipe_list.setEnabled(checked)

    def _browse_path(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Export Save Path")
        if folder:
            self.path_edit.setText(folder)
            self._update_filename()
            self._check_export_ready()

    def _update_filename(self):
        from datetime import datetime
        process = self.process_combo.currentText().strip()
        sheet = self.sheet_combo.currentText().strip()
        if process and sheet:
            date_str = datetime.now().strftime("%y%m%d")
            name = f"{date_str}_{process}_{sheet}.xlsx"
            self.filename_label.setText(f"Filename: {name}")
        else:
            self.filename_label.setText("")

    def _check_export_ready(self):
        path = self.path_edit.text().strip()
        chamber = self.chamber_combo.currentText().strip()
        self.export_btn.setEnabled(bool(path and chamber))

    def _do_export(self):
        """Export 실행"""
        from datetime import datetime
        try:
            import openpyxl
            from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl package is not installed.\npip install openpyxl")
            return

        process = self.process_combo.currentText().strip()
        sheet = self.sheet_combo.currentText().strip()
        chamber = self.chamber_combo.currentText().strip()
        save_dir = self.path_edit.text().strip()

        if not all([process, sheet, chamber, save_dir]):
            QMessageBox.warning(self, "Warning", "Please select all required items.")
            return

        # 선택된 Recipe 코드 목록
        selected_codes = None
        if self.radio_select.isChecked():
            selected_codes = []
            for i in range(self.recipe_list.count()):
                item = self.recipe_list.item(i)
                if item.checkState() == Qt.Checked:
                    selected_codes.append(item.text())
            if not selected_codes:
                QMessageBox.warning(self, "Warning", "Please select at least one Recipe.")
                return

        # 파일 경로
        date_str = datetime.now().strftime("%y%m%d")
        filename = f"{date_str}_{process}_{sheet}.xlsx"
        filepath = os.path.join(save_dir, filename)

        # 파일 덮어쓰기 확인
        if os.path.exists(filepath):
            reply = QMessageBox.question(
                self, "File Exists",
                f"'{filename}' already exists.\nDo you want to overwrite it?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return

        # 데이터 조회
        QApplication.setOverrideCursor(Qt.WaitCursor)
        try:
            db_mgr = self.recipe_service.db_manager
            proc_db = db_mgr.get_process_db_path(process)
            cls_id = db_mgr.get_classification_id(proc_db, sheet, chamber)

            if cls_id is None:
                QApplication.restoreOverrideCursor()
                QMessageBox.warning(self, "Warning", "Classification not found.")
                return

            # code_filter 구성 (특정 Recipe 선택 시)
            code_filter = None
            result = self.recipe_service.load_recipe_data_for_view(
                proc_db, chamber, cls_id, process, code_filter, None
            )

            (dyn_cols, param_ids, rows, groups, base_map, base_lookup,
             recipe_codes, row_stepnos, row_occidx, id_to_name_map,
             dyn_mappings, dyn_units, dense_right_data) = result

            if not rows:
                QApplication.restoreOverrideCursor()
                QMessageBox.information(self, "Info", "No data to export.")
                return

            # 선택된 Recipe만 필터링
            if selected_codes:
                selected_set = set(selected_codes)
                filtered_indices = [i for i, r in enumerate(rows) if r[1][3] in selected_set]
                rows = [rows[i] for i in filtered_indices]
                dense_right_data = [dense_right_data[i] for i in filtered_indices]
                row_occidx = [row_occidx[i] for i in filtered_indices]

                # 그룹 재계산
                groups = []
                if rows:
                    prev_rid = None
                    for i, (rid, _, _) in enumerate(rows):
                        if rid != prev_rid:
                            groups.append({"start": i, "count": 1})
                            prev_rid = rid
                        else:
                            groups[-1]["count"] += 1

            # Excel 생성
            self._generate_excel(
                filepath, rows, groups, dyn_cols, dyn_mappings,
                dense_right_data, base_map, base_lookup, row_occidx
            )

            QApplication.restoreOverrideCursor()
            QMessageBox.information(self, "Success", f"Excel export completed!\n{filepath}")
            self.accept()

        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Error", f"Error during export:\n{e}")

    def _generate_excel(self, filepath, rows, groups, dyn_cols, dyn_mappings,
                        dense_right_data, base_map, base_lookup, row_occidx):
        """UI와 동일한 양식으로 Excel 파일을 생성합니다."""
        import openpyxl
        from openpyxl.styles import Font as XlFont, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recipe"

        # ── 스타일 정의 ──
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = XlFont(bold=True, color="000000", size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        data_align = Alignment(horizontal="right", vertical="center")
        recipe_align = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thick_top = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='thin')
        )

        # ── 헤더 기록 (Row 1) ──
        left_headers = ["Date", "Comment", "Base", "Recipe", "Step"]

        # 줄바꿈 처리 된 컬럼 헤더
        final_cols = [str(col).replace("\\n", "\n") for col in dyn_cols]

        for c, h in enumerate(left_headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        offset = len(left_headers) + 1
        for c, h in enumerate(final_cols):
            cell = ws.cell(row=1, column=offset + c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # ── 데이터 기록 ──
        # 그룹 경계 행 인덱스 (두꺼운 상단 테두리 적용)
        boundary_rows = {g["start"] for g in groups if g.get("start", 0) > 0}

        for r_idx, row_data in enumerate(rows):
            rid, base_vals, param_dict = row_data
            excel_row = r_idx + 2  # 헤더 다음 줄

            is_boundary = (r_idx in boundary_rows)
            border = thick_top if is_boundary else thin_border

            # 왼쪽 컬럼 (Date, Comment, Base, Recipe, Step)
            for c, val in enumerate(base_vals, 1):
                cell = ws.cell(row=excel_row, column=c, value=str(val) if val else "")
                cell.alignment = data_align if c != 4 else recipe_align
                cell.border = border

            # 오른쪽 컬럼 (파라미터)
            dense_row = dense_right_data[r_idx] if r_idx < len(dense_right_data) else []
            for c, val in enumerate(dense_row):
                # 숫자 변환 시도 (Ramp "10 > 20" 등 비숫자는 문자열 유지)
                write_val = ""
                if val:
                    try:
                        write_val = float(val)
                        if write_val == int(write_val):
                            write_val = int(write_val)
                    except (ValueError, TypeError):
                        write_val = val
                cell = ws.cell(row=excel_row, column=offset + c, value=write_val)
                cell.alignment = data_align
                cell.border = border

                # Base 대비 다른 값 → 노란색 배경
                if c < len(dyn_mappings):
                    base_id = base_map.get(rid)
                    if base_id:
                        step_name = base_vals[4]
                        occ_idx = row_occidx[r_idx] if r_idx < len(row_occidx) else 0
                        mapping_key = dyn_mappings[c]
                        current_val = param_dict.get(mapping_key)
                        base_val = base_lookup.get((base_id, step_name, occ_idx, mapping_key))

                        is_diff = False
                        if base_val is not None and current_val is not None:
                            try:
                                if abs(float(base_val) - float(current_val)) > 1e-9:
                                    is_diff = True
                            except (ValueError, TypeError):
                                if str(base_val) != str(current_val):
                                    is_diff = True
                        elif (base_val is None) != (current_val is None):
                            is_diff = True

                        if is_diff:
                            cell.fill = diff_fill

        # ── Recipe 이름 병합 ──
        recipe_col = 4  # Recipe 컬럼 (1-based)
        for g in groups:
            if g["count"] > 1:
                start_row = g["start"] + 2  # Excel row (1-based, 헤더 +1)
                end_row = start_row + g["count"] - 1
                # Date, Comment, Base, Recipe 컬럼 병합
                for merge_col in [1, 2, 3, 4]:
                    ws.merge_cells(
                        start_row=start_row, start_column=merge_col,
                        end_row=end_row, end_column=merge_col
                    )
                    # 병합 셀 중앙 정렬
                    ws.cell(row=start_row, column=merge_col).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

        # ── 컬럼 너비 자동 조정 ──
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 6), 30)

        # ── 틀 고정: 헤더(1행) + 왼쪽 컬럼(Date~Step) 고정 ──
        freeze_col = len(left_headers) + 1  # Step 다음 컬럼 = F
        ws.freeze_panes = f"{get_column_letter(freeze_col)}2"

        wb.save(filepath)